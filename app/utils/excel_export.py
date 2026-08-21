import os
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


DB_NAME = "data/maxky_pos.db"


def create_excel(event_name=""):

    os.makedirs(
        "exports",
        exist_ok=True
    )

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    event_name = event_name.strip()

    # ==========================================
    # ดึงข้อมูลตามชื่องาน
    # ==========================================

    if event_name:

        cursor.execute(
            """
            SELECT
                created_at,
                event_name,
                category,
                size,
                cost,
                price,
                profit

            FROM sales

            WHERE TRIM(event_name) = TRIM(?)

            ORDER BY id DESC
            """,
            (event_name,)
        )

    else:

        cursor.execute(
            """
            SELECT
                created_at,
                event_name,
                category,
                size,
                cost,
                price,
                profit

            FROM sales

            ORDER BY id DESC
            """
        )

    rows = cursor.fetchall()

    # ==========================================
    # สรุปยอด
    # ==========================================

    if event_name:

        cursor.execute(
            """
            SELECT
                COUNT(*),
                COALESCE(SUM(price), 0),
                COALESCE(SUM(cost), 0),
                COALESCE(SUM(profit), 0)

            FROM sales

            WHERE TRIM(event_name) = TRIM(?)
            """,
            (event_name,)
        )

    else:

        cursor.execute(
            """
            SELECT
                COUNT(*),
                COALESCE(SUM(price), 0),
                COALESCE(SUM(cost), 0),
                COALESCE(SUM(profit), 0)

            FROM sales
            """
        )

    summary = cursor.fetchone()

    qty = summary[0] or 0
    sales = summary[1] or 0
    cost = summary[2] or 0
    profit = summary[3] or 0

    conn.close()

    # ==========================================
    # สร้าง Excel
    # ==========================================

    wb = Workbook()

    ws = wb.active
    ws.title = "Sales"

    # ==========================================
    # หัวรายงาน
    # ==========================================

    ws.merge_cells("A1:G1")

    ws["A1"] = "MAXKY POS"

    ws["A1"].font = Font(
        size=20,
        bold=True
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30

    # ==========================================
    # ชื่องาน
    # ==========================================

    ws.merge_cells("A2:G2")

    if event_name:

        ws["A2"] = f"รายงานยอดขาย : {event_name}"

    else:

        ws["A2"] = "รายงานยอดขาย : ทุกงาน"

    ws["A2"].font = Font(
        size=14,
        bold=True
    )

    ws["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[2].height = 25

    # ==========================================
    # สรุปยอด
    # ==========================================

    ws["A4"] = "จำนวนขาย"
    ws["B4"] = qty
    ws["C4"] = "ตัว"

    ws["D4"] = "ยอดขาย"
    ws["E4"] = sales
    ws["F4"] = "บาท"

    ws["A5"] = "ต้นทุน"
    ws["B5"] = cost
    ws["C5"] = "บาท"

    ws["D5"] = "กำไร"
    ws["E5"] = profit
    ws["F5"] = "บาท"

    # ==========================================
    # Style สรุป
    # ==========================================

    summary_fill = PatternFill(
        "solid",
        fgColor="E8F5E9"
    )

    thin = Side(
        style="thin",
        color="D9D9D9"
    )

    summary_border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows(
        min_row=4,
        max_row=5,
        min_col=1,
        max_col=6
    ):

        for cell in row:

            cell.fill = summary_fill

            cell.border = summary_border

            cell.alignment = Alignment(
                vertical="center"
            )

    for cell in [
        "A4",
        "D4",
        "A5",
        "D5"
    ]:

        ws[cell].font = Font(
            bold=True
        )

    for cell in [
        "B4",
        "E4",
        "B5",
        "E5"
    ]:

        ws[cell].font = Font(
            size=13,
            bold=True
        )

    # ==========================================
    # รูปแบบตัวเลข
    # ==========================================

    for cell in [
        "B4",
        "E4",
        "B5",
        "E5"
    ]:

        ws[cell].number_format = '#,##0.00'

    ws["B4"].number_format = '#,##0'
    ws["E4"].number_format = '#,##0.00'
    ws["B5"].number_format = '#,##0.00'
    ws["E5"].number_format = '#,##0.00'

    # ==========================================
    # เว้นบรรทัด
    # ==========================================

    ws["A7"] = "รายละเอียดการขาย"

    ws["A7"].font = Font(
        size=14,
        bold=True
    )

    # ==========================================
    # Header ตาราง
    # ==========================================

    ws.append([])

    ws.append([
        "วันที่",
        "ชื่องาน",
        "สินค้า",
        "ไซส์",
        "ต้นทุน",
        "ราคาขาย",
        "กำไร"
    ])

    header_row = 9

    header_fill = PatternFill(
        "solid",
        fgColor="DDEBF7"
    )

    for cell in ws[header_row]:

        cell.font = Font(
            bold=True
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = summary_border

    # ==========================================
    # Data
    # ==========================================

    for row in rows:

        ws.append(row)

    # ==========================================
    # จัดรูปแบบข้อมูล
    # ==========================================

    for row in ws.iter_rows(
        min_row=10,
        max_row=ws.max_row,
        min_col=1,
        max_col=7
    ):

        for cell in row:

            cell.border = Border(
                bottom=Side(
                    style="hair",
                    color="DDDDDD"
                )
            )

            cell.alignment = Alignment(
                vertical="center"
            )

    # ==========================================
    # รูปแบบเงิน
    # ==========================================

    for row in range(
        10,
        ws.max_row + 1
    ):

        ws[f"E{row}"].number_format = '#,##0.00'
        ws[f"F{row}"].number_format = '#,##0.00'
        ws[f"G{row}"].number_format = '#,##0.00'

    # ==========================================
    # ความกว้างคอลัมน์
    # ==========================================

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # ==========================================
    # Freeze Header
    # ==========================================

    ws.freeze_panes = "A10"

    # ==========================================
    # ชื่อไฟล์
    # ==========================================

    if event_name:

        safe_name = "".join(
            c for c in event_name
            if c not in '\\/:*?"<>|'
        ).strip()

        filename = (
            f"exports/MAXKY_{safe_name}.xlsx"
        )

    else:

        filename = (
            "exports/MAXKY_SALES.xlsx"
        )

    # ==========================================
    # Save
    # ==========================================

    wb.save(filename)

    return filename
